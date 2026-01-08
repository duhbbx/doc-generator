"""Excel file reader module."""

from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReader:
    """Reads data from Excel files."""

    def __init__(self, file_path: str | Path):
        """Initialize the Excel reader.

        Args:
            file_path: Path to the Excel file.
        """
        self.file_path = Path(file_path)
        self._workbook = None
        self._headers: list[str] = []

    def open(self) -> None:
        """Open the Excel file."""
        self._workbook = load_workbook(self.file_path, read_only=True, data_only=True)

    def close(self) -> None:
        """Close the Excel file."""
        if self._workbook:
            self._workbook.close()
            self._workbook = None

    def __enter__(self) -> "ExcelReader":
        self.open()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb) -> None:
        self.close()

    @property
    def sheet_names(self) -> list[str]:
        """Get list of sheet names."""
        if not self._workbook:
            raise RuntimeError("Excel file not opened")
        return self._workbook.sheetnames

    def get_sheet(self, sheet_name: str | None = None) -> Worksheet:
        """Get a worksheet by name.

        Args:
            sheet_name: Name of the sheet. If None, returns the active sheet.

        Returns:
            The worksheet object.
        """
        if not self._workbook:
            raise RuntimeError("Excel file not opened")

        if sheet_name is None:
            return self._workbook.active
        return self._workbook[sheet_name]

    def get_headers(self, sheet_name: str | None = None, header_row: int = 1) -> list[str]:
        """Get column headers from the specified row.

        Args:
            sheet_name: Name of the sheet. If None, uses the active sheet.
            header_row: Row number containing headers (1-indexed).

        Returns:
            List of header names.
        """
        sheet = self.get_sheet(sheet_name)
        headers = []

        for cell in sheet[header_row]:
            value = cell.value
            if value is not None:
                headers.append(str(value))
            else:
                # Stop at first empty cell
                break

        self._headers = headers
        return headers

    def get_column_letters(self, sheet_name: str | None = None, header_row: int = 1) -> dict[str, str]:
        """Get mapping of header names to column letters.

        Args:
            sheet_name: Name of the sheet. If None, uses the active sheet.
            header_row: Row number containing headers (1-indexed).

        Returns:
            Dictionary mapping header names to column letters (A, B, C, etc.)
        """
        sheet = self.get_sheet(sheet_name)
        mapping = {}

        for cell in sheet[header_row]:
            value = cell.value
            if value is not None:
                # Get column letter from cell coordinate
                col_letter = cell.column_letter
                mapping[str(value)] = col_letter
            else:
                break

        return mapping

    def iter_rows(
        self,
        sheet_name: str | None = None,
        header_row: int = 1,
        start_row: int = 2,
    ) -> Iterator[dict[str, any]]:
        """Iterate over data rows, yielding dictionaries.

        Args:
            sheet_name: Name of the sheet. If None, uses the active sheet.
            header_row: Row number containing headers (1-indexed).
            start_row: First row of data (1-indexed).

        Yields:
            Dictionary mapping header names to cell values for each row.
        """
        headers = self.get_headers(sheet_name, header_row)
        sheet = self.get_sheet(sheet_name)

        for row in sheet.iter_rows(min_row=start_row, max_col=len(headers)):
            # Skip completely empty rows
            if all(cell.value is None for cell in row):
                continue

            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = cell.value
            yield row_data

    def get_all_rows(
        self,
        sheet_name: str | None = None,
        header_row: int = 1,
        start_row: int = 2,
    ) -> list[dict[str, any]]:
        """Get all data rows as a list of dictionaries.

        Args:
            sheet_name: Name of the sheet. If None, uses the active sheet.
            header_row: Row number containing headers (1-indexed).
            start_row: First row of data (1-indexed).

        Returns:
            List of dictionaries, each mapping header names to cell values.
        """
        return list(self.iter_rows(sheet_name, header_row, start_row))

    def get_row_count(self, sheet_name: str | None = None, start_row: int = 2) -> int:
        """Get the number of data rows.

        Args:
            sheet_name: Name of the sheet. If None, uses the active sheet.
            start_row: First row of data (1-indexed).

        Returns:
            Number of data rows.
        """
        sheet = self.get_sheet(sheet_name)
        count = 0
        headers = self.get_headers(sheet_name)

        for row in sheet.iter_rows(min_row=start_row, max_col=len(headers)):
            if any(cell.value is not None for cell in row):
                count += 1

        return count
